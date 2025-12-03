# -*- coding: utf-8 -*-
import os
from openpyxl import Workbook, load_workbook

file_path = "nhanvien.xlsx"

# ===============================
# PHẦN 1: Tạo file Excel nếu chưa có
# ===============================
def TaoExcel():
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "NhanVien"
        # tạo header
        ws.append(["STT", "Ma", "Ten", "Tuoi"])
        wb.save(file_path)
        print(f"✅ Đã tạo file Excel mới: {file_path}")
    else:
        print(f"File {file_path} đã tồn tại.")

# ===============================
# PHẦN 2: Đọc dữ liệu Excel
# ===============================
def DocExcel():
    if not os.path.exists(file_path):
        print("❌ File Excel chưa tồn tại!")
        return []
    wb = load_workbook(file_path)
    ws = wb.active
    data = []
    # bỏ header
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is not None:  # Ma
            data.append({"Ma": row[1], "Ten": row[2], "Tuoi": row[3]})
    return data

# ===============================
# PHẦN 3: Lưu dữ liệu Excel
# ===============================
def LuuExcel(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "NhanVien"
    ws.append(["STT", "Ma", "Ten", "Tuoi"])
    for i, nv in enumerate(data, start=1):
        ws.append([i, nv["Ma"], nv["Ten"], nv["Tuoi"]])
    wb.save(file_path)
    print("✅ Dữ liệu đã được lưu vào Excel.")

# ===============================
# PHẦN 4: Chức năng
# ===============================
def XuatDS(data):
    print("\nDanh sách Nhân viên:")
    print("STT\tMa\tTen\tTuoi")
    for i, nv in enumerate(data, start=1):
        print(f"{i}\t{nv['Ma']}\t{nv['Ten']}\t{nv['Tuoi']}")
    print()

def ThemNhanVien(data):
    ma = input("Nhập mã nhân viên: ").strip()
    ten = input("Nhập tên nhân viên: ").strip()
    tuoi = int(input("Nhập tuổi: "))
    data.append({"Ma": ma, "Ten": ten, "Tuoi": tuoi})
    LuuExcel(data)
    print(f"✅ Thêm nhân viên {ma} - {ten} thành công!")

def SapXepTuoi(data):
    data.sort(key=lambda x: x["Tuoi"])
    print("✅ Đã sắp xếp theo tuổi tăng dần.")
    XuatDS(data)

# ===============================
# PHẦN 5: Menu
# ===============================
def Menu():
    TaoExcel()
    data = DocExcel()
    while True:
        print("===== QUẢN LÝ NHÂN VIÊN =====")
        print("1. Thêm nhân viên")
        print("2. Xem danh sách nhân viên")
        print("3. Sắp xếp nhân viên theo tuổi tăng dần")
        print("0. Thoát")
        choice = input("Chọn chức năng: ").strip()
        if choice == "1":
            ThemNhanVien(data)
        elif choice == "2":
            XuatDS(data)
        elif choice == "3":
            SapXepTuoi(data)
        elif choice == "0":
            print("Thoát chương trình.")
            break
        else:
            print("Lựa chọn không hợp lệ.")
        print()

# ===============================
# Chạy chương trình
# ===============================
if __name__ == "__main__":
    Menu()
