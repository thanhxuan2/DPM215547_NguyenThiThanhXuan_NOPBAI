# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import os

# ===============================
# PHẦN 1: Kiểm tra file Excel
# ===============================
path = "demo.xlsx"
if not os.path.exists(path):
    print(f"⚠️ File '{path}' chưa tồn tại. Vui lòng tạo file Excel trước (Câu 7).")
else:
    # ===============================
    # PHẦN 2: Mở workbook và lấy sheet đầu tiên
    # ===============================
    wb = load_workbook(path)
    print("Các sheet trong file:", wb.sheetnames)
    ws = wb[wb.sheetnames[0]]

    # ===============================
    # PHẦN 3: Đọc toàn bộ dữ liệu
    # ===============================
    print("\nDữ liệu trong Excel:")
    for row in ws.values:
        for value in row:
            print(value, "\t", end='')
        print("")  # xuống dòng sau mỗi hàng
